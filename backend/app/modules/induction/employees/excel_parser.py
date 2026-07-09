import datetime
from openpyxl import load_workbook
from app.core.exceptions import InvalidUploadException

def parse_employees_excel(excel_path: str) -> list[dict]:
    """
    Parses employee Excel rows using openpyxl.
    Maps headers dynamically to standardize columns.
    """
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        raise InvalidUploadException(f"Failed to open Excel file: {str(e)}")

    if not wb.sheetnames:
        raise InvalidUploadException("Excel file contains no sheets.")

    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise InvalidUploadException("Excel sheet is empty.")

    # 1. Locate Header Row
    header_idx = -1
    header_mapping = {}

    target_headers = {
        "name": ["name", "employee name", "emp name", "fullname", "full name"],
        "email": ["email", "email address", "mail", "emp email", "email_address"],
        "department": ["department", "dept", "function", "division"],
        "designation": ["designation", "role", "title", "job title", "position"],
        "location": ["location", "office", "city", "site"],
        "joining_date": ["joining date", "join date", "doj", "start date", "joining_date"]
    }

    for r_idx, row in enumerate(rows):
        # Check if this row looks like a header (contains name and email)
        row_str = [str(cell).lower().strip() if cell is not None else "" for cell in row]
        has_name = any(any(h in cell for h in target_headers["name"]) for cell in row_str if cell)
        has_email = any(any(h in cell for h in target_headers["email"]) for cell in row_str if cell)

        if has_name and has_email:
            header_idx = r_idx
            # Map index to target headers
            for c_idx, cell in enumerate(row_str):
                if not cell:
                    continue
                for target, aliases in target_headers.items():
                    if any(alias == cell or alias in cell for alias in aliases):
                        header_mapping[target] = c_idx
            break

    if header_idx == -1:
        # Fallback to row 0 if no header matches
        header_idx = 0
        row_str = [str(cell).lower().strip() if cell is not None else "" for cell in rows[0]]
        for c_idx, cell in enumerate(row_str):
            for target, aliases in target_headers.items():
                if any(alias == cell for alias in aliases):
                    header_mapping[target] = c_idx

    # Require name and email mapping at minimum
    if "name" not in header_mapping or "email" not in header_mapping:
        raise InvalidUploadException("Excel headers must contain at least 'Name' and 'Email' columns.")

    # 2. Extract Data Rows
    employees = []
    for row in rows[header_idx + 1:]:
        # Skip fully empty rows
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        emp_data = {}
        for target, col_idx in header_mapping.items():
            if col_idx < len(row):
                val = row[col_idx]
                # Format datetime values nicely
                if isinstance(val, (datetime.datetime, datetime.date)):
                    val = val.strftime("%Y-%m-%d")
                elif val is not None:
                    val = str(val).strip()
                emp_data[target] = val
            else:
                emp_data[target] = None

        employees.append(emp_data)

    wb.close()
    return employees
